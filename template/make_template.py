#!/usr/bin/env python3
"""旧形式のイベント情報入力シート生成コード（保守停止）。

現在のv2シートは構造が異なるため、このファイルからは再生成しません。
誤って実行しても `template/イベント情報入力シート.xlsx` を上書きしないよう、
実行入口で安全停止します。現行手順は `template/README.md` を参照してください。
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_PATH = Path(__file__).resolve().parent / "イベント情報入力シート.xlsx"

# 共通スタイル
HEADER_FILL = PatternFill("solid", fgColor="F59E0B")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEAD_FILL = PatternFill("solid", fgColor="FDE68A")
SUBHEAD_FONT = Font(bold=True, color="78350F")
NOTE_FONT = Font(italic=True, color="9CA3AF", size=9)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row, values, *, wrap=False):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.alignment = WRAP if wrap else TOP


def write_headers(ws, row, headers, comments=None):
    comments = comments or {}
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        style_header(c)
        if h in comments:
            c.comment = Comment(comments[h], "ひな形")


def build():
    wb = Workbook()

    # ---------------------------------------------------------------
    # シート1「基本情報」(main → eventInfo / events.json)
    # ---------------------------------------------------------------
    ws = wb.active
    ws.title = "基本情報"
    set_widths(ws, [22, 46, 40])

    title = ws.cell(row=1, column=1, value="基本情報（イベント全体の設定）")
    title.font = Font(bold=True, size=13, color="78350F")
    ws.merge_cells("A1:C1")

    write_headers(ws, 2, ["項目", "値（ここに記入）", "補足・記入例"])

    main_rows = [
        ("研究会名", "第37回 全国算数授業研究大会", "JSON: title。チラシ等の正式名称"),
        ("研究会略称", "全国算数授業研究会", "JSON: subtitle / tagline"),
        ("アプリ名（短め）", "全算研2026", "JSON: logoMain。ヘッダー1行目。空欄なら研究会名"),
        ("PWA表示名", "", "JSON: appName（任意）。ホーム画面のアプリ名を更に短くしたいときだけ記入（例: 算サマ2026）。空欄ならアプリ名（logoMain）"),
        ("アイコン3文字", "全算研", "JSON: iconLabel（任意）。アプリアイコンに出る文字。空欄ならアプリ名の先頭3文字"),
        ("ロゴ補足", "全国算数授業研究大会2026", "JSON: logoSub（任意）。ヘッダー2行目は研究会名を優先し、これは予備"),
        ("テーマ", "探究がはじまる！算数で育てたい子どもの「見る目」", "JSON: theme。「――」等で区切るとサブテーマが自動で改行される"),
        ("会場名", "筑波大学附属小学校", "JSON: venueName"),
        ("開催日（表示用）", "令和8年8月4日(火)・5日(水)", "JSON: dateRange。チラシ表記そのまま"),
        ("主催", "全国算数授業研究会", "JSON: organizer（任意）。概要・お知らせの1つ目のカードに表示。複数は「／」区切り"),
        ("共催", "", "JSON: coOrganizer（任意）。ある場合のみ記入。無ければ空欄で行が出ない"),
        ("協力", "", "JSON: cooperation（任意）。ある場合のみ記入"),
        ("後援", "", "JSON: support（任意）。ある場合のみ記入。複数は「／」区切り"),
        ("ブランド色", "#f59e0b", "JSON: brandColor。メイン1色の16進カラー"),
        ("イベントID", "2026-zensanken-37", "ファイル名 events/<id>.json。半角英数・ハイフン・アンダースコアのみ"),
        ("並び順日付", "2026-08-04", "events.json の sortDate。YYYY-MM-DD"),
        ("LINE追加ポップアップ", "false", "JSON: linePromo。LINE友だち追加を促すポップアップを出すか。true=表示／false・空欄=非表示（既定）。チラシQR等で既にLINEから誘導しているイベントは false"),
        ("会場マップ補足", "会場は筑波大学附属小学校です。詳しい会場割りは当日の掲示・配布物をご確認ください。", "JSON: venue.mapNote（任意）"),
    ]
    r = 3
    for k, v, note in main_rows:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v).alignment = WRAP
        nc = ws.cell(row=r, column=3, value=note)
        nc.font = NOTE_FONT
        nc.alignment = WRAP
        r += 1

    # 開催日テーブル（dates[]）
    r += 1
    sub = ws.cell(row=r, column=1, value="■ 開催日ごとの設定（dates[]）")
    sub.font = SUBHEAD_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    write_headers(ws, r, ["ラベル", "日付 / 曜日", "時間"],
                  comments={"ラベル": "例: 1日目 / 2日目"})
    r += 1
    for lab, dw, tm in [
        ("1日目", "2026-08-04 / 火", "10:00〜16:15"),
        ("2日目", "2026-08-05 / 水", "8:50〜16:00"),
    ]:
        write_row(ws, r, [lab, dw, tm])
        r += 1

    # 会場テーブル（rooms[]）
    r += 1
    sub = ws.cell(row=r, column=1, value="■ 会場一覧（rooms[]）※並行セッションがある場合")
    sub.font = SUBHEAD_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    # 色は index.html の COLOR 辞書と対応（blueDeep/greenDeep は同系色の濃色版。
    # 同じ建物の部屋は同系色の濃淡で区別すると会場の位置関係が伝わりやすい）
    write_headers(ws, r, ["会場名", "色（blue/blueDeep/green/greenDeep/orange/purple）", "（空欄可）"])
    r += 1
    for name, color in [("A会場", "blueDeep"), ("B会場", "blue"), ("C会場", "green"), ("D会場", "orange")]:
        write_row(ws, r, [name, color, ""])
        r += 1

    # 現行UIで利用できる任意項目
    r += 1
    sub = ws.cell(row=r, column=1, value="■ 任意設定（必要な場合のみ）")
    sub.font = SUBHEAD_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    optional_rows = [
        ("補助ブランド色", "", "JSON: brandColor2（任意）。ヘッダー等のグラデーション補助色。16進カラー"),
        ("会場マップ画像", "", "JSON: venue.mapImage（任意）。assets/venue-map-<id>.svg 等の相対パスまたは公開URL"),
        ("書籍特設販売URL", "", "JSON: bookSale.url（任意）。BOOKS上部の特設販売ページ導線"),
        ("書籍特設販売ラベル", "", "JSON: bookSale.label（任意）。リンクの表示名"),
        ("書籍特設販売補足", "", "JSON: bookSale.note（任意）。リンク下の短い説明"),
    ]
    for k, v, note in optional_rows:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v).alignment = WRAP
        nc = ws.cell(row=r, column=3, value=note)
        nc.font = NOTE_FONT
        nc.alignment = WRAP
        r += 1

    # ---------------------------------------------------------------
    # シート2「タイムテーブル」(TIMETABLE → sessions)
    # ---------------------------------------------------------------
    ws = wb.create_sheet("タイムテーブル")
    set_widths(ws, [10, 8, 8, 22, 12, 34, 44, 30])
    headers = ["日付ID", "開始", "終了", "区分(category)", "会場(room)",
               "タイトル(title)", "登壇者など(meta)", "補足(note)"]
    comments = {
        "日付ID": "基本情報シートの dates ラベルに対応（day1=1日目, day2=2日目 など）",
        "会場(room)": "全体セッションは空欄。会場別の並行セッションは会場名を入れて行を分ける",
        "タイトル(title)": "受付・昼食など演題が無い区分は空欄でOK",
        "登壇者など(meta)": "1行=1要素。セル内改行（Alt+Enter）で複数行入力すると配列になる",
        "補足(note)": "セッション全体への注記（任意）",
    }
    write_headers(ws, 1, headers, comments)
    ws.freeze_panes = "A2"

    tt_rows = [
        ("day1", "9:30", "10:00", "受付", "", "", "", ""),
        ("day1", "10:00", "10:10", "会長挨拶・諸連絡", "", "", "", ""),
        ("day1", "10:10", "10:40", "基調提案", "",
         "テーマ「子どもを探究へと導く算数の『見る目』」", "提案者：青山尚司（筑波大学附属小）", ""),
        ("day1", "10:50", "11:35", "公開授業1", "",
         "4年「変わり方調べ」", "授業者：森本隆史（筑波大学附属小）\n児童：筑波大学附属小", ""),
        ("day1", "13:30", "14:15", "公開授業2", "A会場",
         "2年「たし算とひき算の筆算」（発展）",
         "授業者：沼川卓也（盛岡市立緑が丘小）\n児童：筑波大学附属小", ""),
        ("day1", "13:30", "14:15", "公開授業2", "B会場",
         "2年「かけ算」",
         "授業者：岡本貴裕（山口市立大内南小）\n児童：筑波大学附属小", ""),
        ("day2", "8:30", "9:00", "受付", "", "", "", ""),
        ("day2", "15:10", "16:00", "講演", "",
         "テーマ「子どもの見る目を育てる」", "講演者：尾崎正彦（国立学園小）", ""),
    ]
    for i, row in enumerate(tt_rows, start=2):
        write_row(ws, i, list(row), wrap=True)

    # ---------------------------------------------------------------
    # シート3「資料リンク」(FILES → resourceLinks / forms)
    # ---------------------------------------------------------------
    ws = wb.create_sheet("資料リンク")
    set_widths(ws, [18, 30, 40, 46])
    headers = ["種別", "ラベル", "説明", "URL"]
    comments = {
        "種別": "「資料」「ホームページ」→ 資料リンク欄 / "
                "「事後アンケート」「参加申し込み」「次回案内希望」→ フォーム欄 に振り分け",
    }
    write_headers(ws, 1, headers, comments)
    ws.freeze_panes = "A2"
    files_rows = [
        ("資料", "当日資料一式（PDF）", "全提案・公開授業の資料フォルダ", "https://drive.google.com/drive/folders/example-all"),
        ("ホームページ", "全国算数授業研究会 ホームページ", "最新情報・お知らせはこちら", "https://zensanken.jimdofree.com/"),
        ("事後アンケート", "事後アンケート（参加者全員）", "本大会の感想をお寄せください", "https://forms.gle/example-survey"),
        ("参加申し込み", "参加お申し込み", "お申し込み時のメールに最新案内をお送りします", "https://www.kokuchpro.com/event/xxxx/"),
        ("次回案内希望", "次回の案内希望", "次回開催や新刊書籍のご案内をLINEでお届けします", "https://lin.ee/xxxx"),
    ]
    for i, row in enumerate(files_rows, start=2):
        write_row(ws, i, list(row), wrap=True)

    # ---------------------------------------------------------------
    # シート4「関連書籍」(BOOKS → books)
    # ---------------------------------------------------------------
    ws = wb.create_sheet("関連書籍")
    set_widths(ws, [34, 30, 16, 40, 60, 40, 36])
    headers = ["書名(title)", "著者(author)", "出版社(publisher)", "表紙URL(cover)",
               "紹介文(description)", "商品URL(url)", "Amazon URL(amazon_url)"]
    write_headers(ws, 1, headers)
    ws.freeze_panes = "A2"
    books_rows = [
        ("多様な子どもが共に学ぶ 学力差に寄り添う算数授業",
         "田中 英海／編著ほか", "東洋館出版社",
         "https://images-na.ssl-images-amazon.com/images/P/4491063540.09.LZZZZZZZ.jpg",
         "学力差を「個性」として活かす、クラス全員に正対した授業づくり。",
         "https://www.toyokan.co.jp/products/6354",
         "https://www.amazon.co.jp/dp/4491063540"),
        ("割合の見方・考え方を働かせる算数授業",
         "青山 尚司／著", "東洋館出版社",
         "https://images-na.ssl-images-amazon.com/images/P/4491063451.09.LZZZZZZZ.jpg",
         "算数の最難単元「割合」を、2つの比べ方で「使える道具」に落とし込む。",
         "https://www.toyokan.co.jp/products/6345",
         "https://www.amazon.co.jp/dp/4491063451"),
    ]
    for i, row in enumerate(books_rows, start=2):
        write_row(ws, i, list(row), wrap=True)

    # ---------------------------------------------------------------
    # シート5「概要・お知らせ」(notices)
    # ---------------------------------------------------------------
    ws = wb.create_sheet("概要・お知らせ")
    set_widths(ws, [34, 80, 22])
    headers = ["タイトル(title)", "本文(body)", "重要度(level)"]
    comments = {"重要度(level)": "important（強調表示）または info（通常）"}
    write_headers(ws, 1, headers, comments)
    ws.freeze_panes = "A2"
    notice_rows = [
        ("ご来場にあたって（体調管理のお願い）",
         "発熱・咳などの症状がある方は参加を見合わせてください。会場では手指消毒・換気にご協力をお願いします。",
         "important"),
        ("テーマ実践書籍の特別販売",
         "大会テーマに合わせた実践書籍を会場で特別販売します。ぜひお立ち寄りください。",
         "info"),
    ]
    for i, row in enumerate(notice_rows, start=2):
        write_row(ws, i, list(row), wrap=True)

    wb.save(OUT_PATH)
    print(f"生成しました: {OUT_PATH}")


if __name__ == "__main__":
    raise SystemExit(
        "このスクリプトは旧形式のため実行できません。"
        "現行の template/イベント情報入力シート.xlsx を使用してください。"
    )

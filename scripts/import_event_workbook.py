#!/usr/bin/env python3
"""v2イベント入力シートを検証し、イベントファイルを決定的に生成する。

既定は読み取り専用の事前検証。`--write` を付けた場合だけリポジトリへ書く。

使い方:
    python scripts/import_event_workbook.py path/to/event.xlsx
    python scripts/import_event_workbook.py path/to/event.xlsx --write
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from collections import defaultdict
from copy import deepcopy
from dataclasses import dataclass
from datetime import date, datetime, time
from html import escape
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from generate_event_url_index import write_event_url_index

try:  # Windows の cp932 端末でも UTF-8 で出力する
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - 利用環境向け案内
    raise SystemExit(
        "openpyxl が必要です。python -m pip install -r requirements-tools.txt を実行してください。"
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
REQUIRED_SHEETS = {
    "入力ガイド",
    "参照元",
    "基本情報",
    "開催日",
    "会場",
    "セッション",
    "プログラム項目",
    "資料リンク",
    "関連書籍",
    "概要・お知らせ",
    "要確認",
}
ID_RE = re.compile(r"^[A-Za-z0-9_-]+$")
HEX_RE = re.compile(r"^#[0-9A-Fa-f]{6}$")
TIME_RE = re.compile(r"^(?:[01]?\d|2[0-3]):[0-5]\d$")
ROOM_COLORS = {"blue", "blueDeep", "green", "greenDeep", "orange", "purple"}
LINK_TYPES = {
    "資料": "resource",
    "ホームページ": "resource",
    "事後アンケート": "form",
    "参加申し込み": "form",
    "次回案内希望": "form",
}
PLACEHOLDER_MARKERS = ("example", "xxxx", "<id>")


@dataclass
class Issue:
    level: str
    location: str
    message: str


class WorkbookError(Exception):
    """入力シートにエラーがある。"""


def relative_luminance(hex_color: str) -> float:
    """#RRGGBB の相対輝度（WCAG 定義）。index.html の relativeLuminance() と同じ式・同じ閾値を使う。"""
    h = hex_color.strip().lstrip("#")
    r, g, b = (int(h[i : i + 2], 16) for i in (0, 2, 4))

    def lin(v: int) -> float:
        v /= 255
        return v / 12.92 if v <= 0.03928 else ((v + 0.055) / 1.055) ** 2.4

    return 0.2126 * lin(r) + 0.7152 * lin(g) + 0.0722 * lin(b)


def icon_fg_color(brand_color: str) -> str:
    """アイコン背景色に対して読みやすい文字色（濃色/白）を選ぶ。index.html の updateBrandFg() と同じ閾値 0.55。"""
    return "#1e293b" if relative_luminance(brand_color) > 0.55 else "#ffffff"


def text(value: Any) -> str:
    """Excel値を前後空白なしの文字列へ変換する。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def split_lines(value: Any) -> list[str]:
    """セル内改行を空行なしの配列へ変換する。"""
    return [line.strip() for line in text(value).replace("\r", "").split("\n") if line.strip()]


def parse_bool(value: Any) -> bool | None:
    raw = text(value).lower()
    if raw in {"true", "1", "yes"}:
        return True
    if raw in {"false", "0", "no", ""}:
        return False
    return None


def parse_order(value: Any) -> int | None:
    raw = text(value)
    if not raw:
        return None
    try:
        parsed = int(float(raw))
    except ValueError:
        return None
    return parsed if parsed > 0 else None


def table_rows(sheet: Any, header_row: int = 3) -> list[dict[str, Any]]:
    """見出し名をキーとして、空行を除く表データを返す。"""
    headers = [text(cell.value) for cell in sheet[header_row]]
    rows: list[dict[str, Any]] = []
    iterator = sheet.iter_rows(min_row=header_row + 1)
    for row_number, cells in enumerate(iterator, start=header_row + 1):
        values = [cell.value for cell in cells[: len(headers)]]
        if not any(text(value) for value in values):
            continue
        record = {headers[i]: values[i] for i in range(len(headers)) if headers[i]}
        record["__row__"] = row_number
        rows.append(record)
    return rows


def atomic_json(path: Path, value: Any) -> None:
    """UTF-8 JSONを同一ディレクトリ内で原子的に置換する。"""
    payload = json.dumps(value, ensure_ascii=False, indent=2) + "\n"
    path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(
        "w", encoding="utf-8", newline="\n", dir=path.parent, delete=False
    ) as tmp:
        tmp.write(payload)
        temp_path = Path(tmp.name)
    temp_path.replace(path)


class EventWorkbook:
    """v2入力シートの読み取り・検証・変換を担当する。"""

    def __init__(self, path: Path) -> None:
        self.path = path
        self.issues: list[Issue] = []
        self.workbook = load_workbook(path, data_only=True, read_only=True)
        missing = sorted(REQUIRED_SHEETS - set(self.workbook.sheetnames))
        for name in missing:
            self.error("workbook", f"必須シートがありません: {name}")

    def error(self, location: str, message: str) -> None:
        self.issues.append(Issue("ERROR", location, message))

    def warn(self, location: str, message: str) -> None:
        self.issues.append(Issue("WARN", location, message))

    def sheet(self, name: str) -> Any:
        if name not in self.workbook.sheetnames:
            raise WorkbookError(f"必須シートがありません: {name}")
        return self.workbook[name]

    def basic_values(self) -> dict[str, str]:
        values: dict[str, str] = {}
        sheet = self.sheet("基本情報")
        for cells in sheet.iter_rows(min_row=4, max_col=2):
            key = text(cells[0].value)
            if key:
                values[key] = text(cells[1].value)
        return values

    def validate(self) -> None:
        if any(issue.level == "ERROR" for issue in self.issues):
            return
        basic = self.basic_values()
        required = (
            "研究会名",
            "アプリ名（短め）",
            "テーマ",
            "会場名",
            "開催日（表示用）",
            "ブランド色",
            "イベントID",
            "並び順日付",
        )
        for label in required:
            if not basic.get(label):
                self.error(f"基本情報.{label}", "必須項目です")
        event_id = basic.get("イベントID", "")
        if event_id and not ID_RE.fullmatch(event_id):
            self.error("基本情報.イベントID", "半角英数・ハイフン・アンダースコアのみ使用できます")
        for label in ("ブランド色", "補助ブランド色"):
            value = basic.get(label, "")
            if value and not HEX_RE.fullmatch(value):
                self.error(f"基本情報.{label}", "# + 6桁16進で指定してください")
        try:
            date.fromisoformat(basic.get("並び順日付", ""))
        except ValueError:
            self.error("基本情報.並び順日付", "YYYY-MM-DDの実在日付が必要です")
        if parse_bool(basic.get("LINE追加ポップアップ")) is None:
            self.error("基本情報.LINE追加ポップアップ", "true または false が必要です")

        sources = table_rows(self.sheet("参照元"))
        if not sources:
            self.error("参照元", "AIへ渡した資料を1件以上記録してください")
        for row in sources:
            loc = f"参照元.{row['__row__']}行"
            if text(row.get("優先度")) not in {"1（最優先）", "2", "3"}:
                self.error(f"{loc}.優先度", "1（最優先）/ 2 / 3 のいずれかが必要です")
            if text(row.get("種別")) not in {"申込ページ", "チラシ", "プログラム", "公式サイト", "その他"}:
                self.error(f"{loc}.種別", "ドロップダウンの値を使用してください")
            if not text(row.get("資料名")):
                self.error(f"{loc}.資料名", "必須項目です")
            if not text(row.get("URLまたはファイル名")):
                self.error(f"{loc}.URLまたはファイル名", "必須項目です")

        dates = table_rows(self.sheet("開催日"))
        date_ids = self.unique_column(dates, "日付ID", "開催日")
        if not date_ids:
            self.error("開催日", "1行以上必要です")
        for row in dates:
            loc = f"開催日.{row['__row__']}行"
            for key in ("日付ID", "表示ラベル", "日付", "曜日", "開催時間（表示用）"):
                if not text(row.get(key)):
                    self.error(f"{loc}.{key}", "必須項目です")
            try:
                date.fromisoformat(text(row.get("日付")))
            except ValueError:
                self.error(f"{loc}.日付", "YYYY-MM-DDの実在日付が必要です")

        rooms = table_rows(self.sheet("会場"))
        room_ids = self.unique_column(rooms, "会場ID", "会場")
        for row in rooms:
            loc = f"会場.{row['__row__']}行"
            if not text(row.get("画面の表示名")):
                self.error(f"{loc}.画面の表示名", "会場IDを入力した行では必須です")
            color = text(row.get("色"))
            if color not in ROOM_COLORS:
                self.error(f"{loc}.色", f"許可値: {', '.join(sorted(ROOM_COLORS))}")

        sessions = table_rows(self.sheet("セッション"))
        session_ids = self.unique_column(sessions, "セッションキー", "セッション")
        if not session_ids:
            self.error("セッション", "1行以上必要です")
        for row in sessions:
            loc = f"セッション.{row['__row__']}行"
            date_id = text(row.get("日付ID"))
            if date_id not in date_ids:
                self.error(f"{loc}.日付ID", "開催日シートに存在しません")
            if not text(row.get("区分")):
                self.error(f"{loc}.区分", "必須項目です")
            for key in ("開始", "終了"):
                value = text(row.get(key))
                if value and not TIME_RE.fullmatch(value):
                    self.error(f"{loc}.{key}", "H:MM または HH:MM で入力してください")

        items = table_rows(self.sheet("プログラム項目"))
        seen_orders: set[tuple[str, int]] = set()
        used_sessions: set[str] = set()
        for row in items:
            loc = f"プログラム項目.{row['__row__']}行"
            session_id = text(row.get("セッションキー"))
            if session_id not in session_ids:
                self.error(f"{loc}.セッションキー", "セッションシートに存在しません")
            else:
                used_sessions.add(session_id)
            order = parse_order(row.get("項目順"))
            if order is None:
                self.error(f"{loc}.項目順", "1以上の整数が必要です")
            elif (session_id, order) in seen_orders:
                self.error(f"{loc}.項目順", "同じセッション内で重複しています")
            else:
                seen_orders.add((session_id, order))
            room_id = text(row.get("会場ID"))
            if room_id and room_id not in room_ids:
                self.error(f"{loc}.会場ID", "会場シートに存在しません")
            if not text(row.get("タイトル")) and not split_lines(row.get("登壇者など（1行1要素）")):
                self.error(loc, "タイトルまたは登壇者などのどちらかが必要です")
            if parse_bool(row.get("控えめ表示")) is None:
                self.error(f"{loc}.控えめ表示", "true / false / 空欄のいずれかが必要です")

        unresolved = [
            row for row in table_rows(self.sheet("要確認"))
            if text(row.get("状態")) != "確認済み"
        ]
        for row in unresolved:
            self.error(f"要確認.{row['__row__']}行", "未確認事項が残っています")

        for row in table_rows(self.sheet("資料リンク")):
            loc = f"資料リンク.{row['__row__']}行"
            if text(row.get("種別")) not in LINK_TYPES:
                self.error(f"{loc}.種別", "ドロップダウンの値を使用してください")
            for key in ("ラベル", "公開URL"):
                if not text(row.get(key)):
                    self.error(f"{loc}.{key}", "必須項目です")

        for row in table_rows(self.sheet("関連書籍")):
            loc = f"関連書籍.{row['__row__']}行"
            for key in ("書名", "著者", "出版社", "紹介文", "商品URL"):
                if not text(row.get(key)):
                    self.error(f"{loc}.{key}", "必須項目です")
            for key in ("表紙URL", "商品URL", "Amazon URL"):
                url = text(row.get(key))
                if url and not url.startswith(("http://", "https://")):
                    self.error(f"{loc}.{key}", "http(s)の公開URLが必要です")

        for row in table_rows(self.sheet("概要・お知らせ")):
            loc = f"概要・お知らせ.{row['__row__']}行"
            for key in ("タイトル", "本文"):
                if not text(row.get(key)):
                    self.error(f"{loc}.{key}", "必須項目です")
            if text(row.get("重要度")) not in {"important", "info"}:
                self.error(f"{loc}.重要度", "important または info が必要です")

        for sheet_name, url_header in (
            ("資料リンク", "公開URL"),
            ("関連書籍", "商品URL"),
        ):
            for row in table_rows(self.sheet(sheet_name)):
                url = text(row.get(url_header))
                if not url.startswith(("http://", "https://")):
                    self.error(f"{sheet_name}.{row['__row__']}行.{url_header}", "公開URLが必要です")
                elif any(marker in url.lower() for marker in PLACEHOLDER_MARKERS):
                    self.error(f"{sheet_name}.{row['__row__']}行.{url_header}", "仮URLを公開できません")

    def unique_column(self, rows: list[dict[str, Any]], key: str, sheet_name: str) -> set[str]:
        result: set[str] = set()
        for row in rows:
            value = text(row.get(key))
            if not value:
                self.error(f"{sheet_name}.{row['__row__']}行.{key}", "必須項目です")
            elif value in result:
                self.error(f"{sheet_name}.{row['__row__']}行.{key}", "重複しています")
            result.add(value)
        return result

    def build(self) -> tuple[str, dict[str, Any], dict[str, Any], dict[str, Any], str]:
        """検証済みExcelから生成対象を返す。"""
        if any(issue.level == "ERROR" for issue in self.issues):
            raise WorkbookError("入力エラーがあるため変換できません")
        basic = self.basic_values()
        event_id = basic["イベントID"]

        info: dict[str, Any] = {
            "title": basic["研究会名"],
        }
        optional_simple = (
            ("研究会略称", "subtitle"),
            ("アプリ名（短め）", "logoMain"),
            ("ロゴ補足", "logoSub"),
            ("PWA表示名", "appName"),
            ("アイコン3文字", "iconLabel"),
            ("ブランド色", "brandColor"),
            ("補助ブランド色", "brandColor2"),
            ("テーマ", "theme"),
            ("会場名", "venueName"),
            ("開催日（表示用）", "dateRange"),
        )
        for label, key in optional_simple:
            if basic.get(label):
                info[key] = basic[label]
        info["manifestPath"] = f"events/{event_id}.webmanifest"
        for label, key in (
            ("主催", "organizer"),
            ("共催", "coOrganizer"),
            ("協力", "cooperation"),
            ("後援", "support"),
        ):
            organizations = split_lines(basic.get(label))
            if organizations:
                info[key] = organizations[0] if len(organizations) == 1 else organizations
        if basic.get("研究会略称"):
            info["tagline"] = basic["研究会略称"]
        info["linePromo"] = bool(parse_bool(basic.get("LINE追加ポップアップ")))

        info["dates"] = [
            {
                "id": text(row["日付ID"]),
                "label": text(row["表示ラベル"]),
                "date": text(row["日付"]),
                "weekday": text(row["曜日"]),
                "time": text(row["開催時間（表示用）"]),
            }
            for row in table_rows(self.sheet("開催日"))
        ]
        info["notices"] = [
            {
                "title": text(row["タイトル"]),
                "body": text(row["本文"]),
                "level": text(row["重要度"]),
            }
            for row in table_rows(self.sheet("概要・お知らせ"))
        ]

        resource_links: list[dict[str, str]] = []
        forms: list[dict[str, str]] = []
        for row in table_rows(self.sheet("資料リンク")):
            link = {
                "label": text(row["ラベル"]),
                "description": text(row["説明"]),
                "url": text(row["公開URL"]),
            }
            target = LINK_TYPES.get(text(row["種別"]))
            (resource_links if target == "resource" else forms).append(link)
        info["forms"] = forms

        book_sale = {
            "url": basic.get("書籍特設販売URL", ""),
            "label": basic.get("書籍特設販売ラベル", ""),
            "note": basic.get("書籍特設販売補足", ""),
        }
        if book_sale["url"]:
            info["bookSale"] = {key: value for key, value in book_sale.items() if value}

        venue: dict[str, Any] = {}
        if basic.get("会場マップ画像"):
            venue["mapImage"] = basic["会場マップ画像"]
        if basic.get("会場マップ補足"):
            venue["mapNote"] = basic["会場マップ補足"]
        venue["resourceLinks"] = resource_links
        info["venue"] = venue

        rooms = [
            {"id": text(row["会場ID"]), "name": text(row["画面の表示名"]), "color": text(row["色"])}
            for row in table_rows(self.sheet("会場"))
        ]
        items_by_session: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
        for row in table_rows(self.sheet("プログラム項目")):
            item: dict[str, Any] = {}
            if text(row.get("会場ID")):
                item["room"] = text(row["会場ID"])
            if text(row.get("タイトル")):
                item["title"] = text(row["タイトル"])
            meta = split_lines(row.get("登壇者など（1行1要素）"))
            if meta:
                item["meta"] = meta
            if parse_bool(row.get("控えめ表示")):
                item["subtle"] = True
            items_by_session[text(row["セッションキー"])].append((parse_order(row["項目順"]) or 0, item))

        sessions: list[dict[str, Any]] = []
        for row in table_rows(self.sheet("セッション")):
            session_id = text(row["セッションキー"])
            session: dict[str, Any] = {
                "id": session_id,
                "dateId": text(row["日付ID"]),
                "start": text(row.get("開始")),
                "end": text(row.get("終了")),
                "category": text(row["区分"]),
            }
            if text(row.get("セッション補足")):
                session["note"] = text(row["セッション補足"])
            if text(row.get("セッション後補足")):
                session["afterNote"] = text(row["セッション後補足"])
            ordered_items = sorted(
                items_by_session[session_id], key=lambda pair: pair[0]
            )
            session["items"] = [item for _, item in ordered_items]
            sessions.append(session)

        books: list[dict[str, Any]] = []
        for row in table_rows(self.sheet("関連書籍")):
            book = {
                "title": text(row["書名"]),
                "author": text(row["著者"]),
                "publisher": text(row["出版社"]),
                "cover": text(row["表紙URL"]),
                "description": text(row["紹介文"]),
                "url": text(row["商品URL"]),
                "amazon_url": text(row["Amazon URL"]),
            }
            books.append({key: value for key, value in book.items() if value})

        event_data = {"id": event_id, "eventInfo": info, "rooms": rooms, "sessions": sessions, "books": books}
        index_entry = {
            "id": event_id,
            "title": info["title"],
            "theme": info["theme"],
            "dateRange": info["dateRange"],
            "venueName": info["venueName"],
            "sortDate": basic["並び順日付"],
            "brandColor": info["brandColor"],
        }
        app_name = info.get("appName") or info["logoMain"]
        icon_label = info.get("iconLabel") or info["logoMain"][:3]
        manifest = {
            "name": app_name,
            "short_name": app_name,
            "description": f"{info['title']}のプログラム・配布資料・関連書籍",
            "lang": "ja",
            "start_url": f"../?id={event_id}",
            "scope": "../",
            "display": "standalone",
            "orientation": "portrait",
            "background_color": "#f8fafc",
            "theme_color": info["brandColor"],
            "icons": [
                {
                    "src": f"../assets/icon-{event_id}.svg",
                    "sizes": "any",
                    "type": "image/svg+xml",
                    "purpose": "any",
                },
                {
                    "src": f"../assets/icon-{event_id}.svg",
                    "sizes": "any",
                    "type": "image/svg+xml",
                    "purpose": "maskable",
                },
            ],
        }
        icon_svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 512 512" '
            'width="512" height="512">\n'
            f'  <rect width="512" height="512" rx="96" fill="{info["brandColor"]}"/>\n'
            f'  <g fill="{icon_fg_color(info["brandColor"])}" font-family="\'Hiragino Sans\',\'Yu Gothic\','
            '\'Noto Sans JP\',sans-serif">\n'
            '    <text x="256" y="256" font-size="130" font-weight="800" '
            'text-anchor="middle" dominant-baseline="central">'
            f"{escape(icon_label)}</text>\n"
            "  </g>\n</svg>\n"
        )
        return event_id, event_data, index_entry, manifest, icon_svg


def write_outputs(
    event_id: str,
    event_data: dict[str, Any],
    index_entry: dict[str, Any],
    manifest: dict[str, Any],
    icon_svg: str,
) -> None:
    """既存イベントを保護しながら生成物を書く。"""
    event_path = ROOT / "events" / f"{event_id}.json"
    if event_path.exists():
        raise WorkbookError(f"既存イベントは上書きできません: {event_path.relative_to(ROOT)}")
    index_path = ROOT / "events.json"
    index_data = json.loads(index_path.read_text(encoding="utf-8"))
    entries = deepcopy(index_data.get("events", []))
    if any(entry.get("id") == event_id for entry in entries):
        raise WorkbookError(f"events.json に同じIDがあります: {event_id}")
    entries.append(index_entry)
    entries.sort(key=lambda entry: (entry.get("sortDate", ""), entry.get("id", "")), reverse=True)
    index_data["events"] = entries

    atomic_json(event_path, event_data)
    atomic_json(ROOT / "events" / f"{event_id}.webmanifest", manifest)
    (ROOT / "assets" / f"icon-{event_id}.svg").write_text(icon_svg, encoding="utf-8", newline="\n")
    atomic_json(index_path, index_data)
    write_event_url_index(index_data)

    result = subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "validate_events.py"), "--event", event_id],
        cwd=ROOT,
        check=False,
    )
    if result.returncode:
        raise WorkbookError("生成後の検証に失敗しました。git diffを確認してください")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="入力済みv2 Excel")
    parser.add_argument("--write", action="store_true", help="検証成功後にリポジトリへ生成物を書く")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.workbook.is_file():
        print(f"ERROR: ファイルがありません: {args.workbook}")
        return 1
    source = EventWorkbook(args.workbook)
    source.validate()
    for issue in source.issues:
        print(f"{issue.level}: {issue.location}: {issue.message}")
    error_count = sum(issue.level == "ERROR" for issue in source.issues)
    warning_count = sum(issue.level == "WARN" for issue in source.issues)
    print(f"Excel事前検証: エラー{error_count}件 / 警告{warning_count}件")
    if error_count:
        return 1
    event_id, event_data, index_entry, manifest, icon_svg = source.build()
    if not args.write:
        print(f"変換準備OK: {event_id}（書き込みなし。生成するには --write を追加）")
        return 0
    try:
        write_outputs(event_id, event_data, index_entry, manifest, icon_svg)
    except WorkbookError as exc:
        print(f"ERROR: {exc}")
        return 1
    print(f"生成完了: events/{event_id}.json と関連3ファイル")
    return 0


if __name__ == "__main__":
    sys.exit(main())
